"""Genera la definición PHP del Índice de Concentración Turística directamente
desde el instrumento original, para evitar errores de transcripción.

A diferencia de las otras matrices generadas, esta no puntúa criterios en una
escala: cuenta cosas. El instrumento trae 113 subtipos/subcategorías y el
riesgo no es la lógica sino los nombres: un nombre de campo mal copiado no
rompe ningún test, solo desvía un conteo en silencio. Por eso se generan desde
el Excel en vez de escribirse a mano, igual que se hizo con Valoración
Territorial.

Uso, desde la raíz del proyecto:  python database/matrices/generar_concentracion.py
"""
import re
import unicodedata
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[2]
ORIGEN = RAIZ / "Documentación" / "Índice de Concentración Turística.xlsx"
DESTINO = RAIZ / "app" / "Matrices" / "Concentracion.php"

wb = openpyxl.load_workbook(ORIGEN, data_only=True)


def limpiar(t):
    """Colapsa espacios y saltos de línea internos. Varias celdas de este
    instrumento traen el nombre partido en dos líneas dentro de la misma
    celda (p. ej. 'Ambientes\\nLacústres', 'Alojamiento\\n(AL)')."""
    return " ".join(str(t).split()) if t else ""


# PostgreSQL trunca en silencio cualquier identificador de más de 63 bytes
# (NAMEDATALEN): dos columnas que solo difieren después del byte 63
# quedarían indistinguibles en producción aunque en SQLite (desarrollo y
# tests) convivan sin problema -exactamente el patrón de fallo que este
# proyecto ya sufrió por partida doble: la suite en verde y la base real
# haciendo otra cosa-. Por eso el límite se aplica aquí, al construir el
# nombre, y no se deja como algo que un test descubra más tarde.
#
# Este mapa recoge las palabras largas y recurrentes de la taxonomía del
# instrumento -"realizaciones", "explotaciones", "acontecimientos"...-, para
# que el generador pueda acortar un nombre sin inventar una abreviatura
# distinta cada vez. Es una decisión legible a propósito: si el instrumento
# cambia y aparece una palabra nueva que empuja algún nombre por encima de
# 63 bytes, `registrar_en_mapa` aborta señalando cuál, y quien lo vea sabrá
# que la solución es añadir una entrada aquí, no tocar el límite.
LIMITE_IDENTIFICADOR_POSTGRES = 63

ABREVIATURAS = {
    "realizaciones": "realizac",
    "tecnicas": "tec",
    "cientificas": "cientif",
    "explotaciones": "explotac",
    "agropecuarias": "agropec",
    "pesqueras": "pesq",
    "exhibicion": "exhib",
    "acontecimientos": "aconteci",
    "programados": "program",
    "convenciones": "convenc",
    "artesanales": "artesan",
    "congresos": "congr",
    "industriales": "industr",
    "religiosas": "relig",
    "tradicionales": "tradic",
    "creencias": "creenc",
    "populares": "popul",
}


def slug(t):
    """Nombre de campo a partir de un texto del instrumento: minúsculas, sin
    acentos, sin puntuación, con las palabras largas de ABREVIATURAS ya
    acortadas. `unicodedata` + encode/decode ascii es más robusto que una
    tabla de reemplazos á->a manual porque cubre cualquier acento o diéresis
    sin tener que enumerarlos.

    Se abrevia palabra por palabra -no el nombre completo- para que el mismo
    término se acorte siempre igual sin importar en qué campo aparezca."""
    t = limpiar(t).lower()
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode("ascii")
    palabras = re.findall(r"[a-z0-9]+", t)
    palabras = [ABREVIATURAS.get(p, p) for p in palabras]
    return "_".join(palabras)


def leer_atractivos(hoja, col_cat, col_tipo, col_subtipo, fila_ini, fila_fin_esperada, bloque_sigla, vistos):
    """Lee una de las dos tablas de la hoja de atractivos (manifestaciones
    culturales o atractivos naturales). CATEGORÍA y TIPO están en celdas
    combinadas -solo traen valor en la fila donde empieza el grupo-, así que
    se arrastran hacia abajo.

    El árbol resultante es categoría -> tipo -> {campo: subtipo}: el tipo
    (Arquitectura, Montaña, Folklore...) es el nivel de agrupación útil para
    una interfaz -entre 3 y 8 subtipos por tipo-, mientras que la categoría
    por sí sola metería 22 o 55 campos seguidos en una sola sección.

    El final de la tabla se detecta por contenido (la fila cuya columna de
    categoría empieza con 'TOTAL'), no por un número de fila fijo: los rangos
    que trae el plan salen de una inspección manual y podrían estar mal. Aun
    así, se contrasta el rango detectado contra ese rango esperado y se
    aborta si no coincide, para que un instrumento distinto al que se validó
    a mano no pase inadvertido.
    """
    categoria_actual = None
    tipo_actual = None
    grupo = {}  # categoría -> tipo -> {campo: etiqueta}, en orden de aparición
    fila = fila_ini
    while True:
        v_cat = hoja.cell(fila, col_cat).value
        v_tipo = hoja.cell(fila, col_tipo).value
        v_subtipo = hoja.cell(fila, col_subtipo).value

        if v_cat is not None:
            texto_cat = limpiar(v_cat)
            if texto_cat.upper().startswith("TOTAL"):
                break
            categoria_actual = texto_cat
        if v_tipo is not None:
            tipo_actual = limpiar(v_tipo)

        if categoria_actual is None or tipo_actual is None:
            raise ValueError(f"'{hoja.title}' fila {fila}: falta categoría o tipo antes de esta fila")

        subtipo = limpiar(v_subtipo)
        if not subtipo:
            raise ValueError(f"'{hoja.title}' fila {fila}: subtipo vacío (categoría={categoria_actual!r})")

        campo = f"at_{bloque_sigla}_{slug(tipo_actual)}_{slug(subtipo)}"
        contexto = f"'{hoja.title}' fila {fila} ({categoria_actual} / {tipo_actual} / {subtipo})"

        # Título con capitalización de presentación: el instrumento trae la
        # categoría en mayúsculas sostenidas ('MANIFESTACIONES CULTURALES'),
        # que aquí se usa como bloque de primer nivel en la interfaz.
        titulo_categoria = categoria_actual.title()
        grupo.setdefault(titulo_categoria, {})
        grupo[titulo_categoria].setdefault(tipo_actual, {})
        # La etiqueta es solo el subtipo: el tipo ya es la clave del nivel
        # que lo contiene, repetirlo en cada etiqueta sería ruido.
        registrar_en_mapa(grupo[titulo_categoria][tipo_actual], campo, subtipo, vistos, contexto)
        fila += 1

    fila_fin_detectada = fila - 1
    if (fila_ini, fila_fin_detectada) != (fila_ini, fila_fin_esperada):
        raise ValueError(
            f"'{hoja.title}': la tabla que empieza en {fila_ini} termina en la fila "
            f"{fila_fin_detectada} (total en {fila}), no en la fila {fila_fin_esperada} "
            "que se había verificado a mano. Revisa el instrumento antes de continuar."
        )
    return grupo


def registrar_en_mapa(mapa_campo_etiqueta, campo, etiqueta, vistos, contexto):
    """Añade `campo` -> `etiqueta` a `mapa_campo_etiqueta`, o aborta.

    Dos invariantes, en este orden:

    1. El nombre cabe en un identificador de PostgreSQL. Abreviar acerca
       nombres que antes eran bien distintos, así que esta comprobación
       importa más que antes de que existiera ABREVIATURAS: hay que saber
       si sigue haciendo falta una abreviatura más, no descubrirlo en
       producción.
    2. El nombre no se repite. Es el fallo más probable de esta taxonomía
       -algún subtipo se repite bajo dos tipos distintos, o el acortamiento
       de la invariante anterior hace coincidir a dos que antes diferían-,
       y hay que pararlo aquí, no resolverlo por su cuenta: dos filas con el
       mismo campo escribirían en la misma columna y una pisaría a la otra
       en silencio.
    """
    if len(campo) > LIMITE_IDENTIFICADOR_POSTGRES:
        raise ValueError(
            f"nombre de campo de {len(campo)} caracteres, por encima del límite de "
            f"{LIMITE_IDENTIFICADOR_POSTGRES} de PostgreSQL: {campo!r}\n"
            f"  producido por: {contexto}\n"
            "  añade una abreviatura para su(s) palabra(s) larga(s) al mapa ABREVIATURAS."
        )
    if campo in vistos:
        raise ValueError(
            f"nombre de campo repetido: {campo!r}\n"
            f"  ya asignado a: {vistos[campo]}\n"
            f"  también producido por: {contexto}"
        )
    vistos[campo] = contexto
    mapa_campo_etiqueta[campo] = etiqueta


def leer_planta(hoja, col_sector, col_categoria, fila_ini, fila_fin_esperada, vistos):
    """Lee la hoja de planta turística. El sector está en celdas combinadas
    -solo trae valor en su primera fila- y arrastra su sigla entre paréntesis
    ('Alojamiento (AL)') a todas las filas de su grupo; esa sigla, en
    minúsculas, es el prefijo del campo.

    Igual que en atractivos, el final de la tabla se detecta por la fila de
    total y se contrasta contra el rango verificado a mano.
    """
    sector_actual = None
    sigla_actual = None
    planta = {}  # sector -> {campo: etiqueta}, en orden de aparición
    fila = fila_ini
    while True:
        v_sector = hoja.cell(fila, col_sector).value
        v_categoria = hoja.cell(fila, col_categoria).value

        if v_sector is not None:
            texto_sector = limpiar(v_sector)
            if texto_sector.upper().startswith("TOTAL"):
                break
            coincidencia = re.search(r"\(([A-ZÁÉÍÓÚÑ]+)\)\.?\s*$", texto_sector)
            if not coincidencia:
                raise ValueError(
                    f"'{hoja.title}' fila {fila}: el sector {texto_sector!r} no trae una "
                    "sigla entre paréntesis al final"
                )
            sector_actual = texto_sector
            sigla_actual = coincidencia.group(1).lower()

        if sector_actual is None:
            raise ValueError(f"'{hoja.title}' fila {fila}: falta sector antes de esta fila")

        categoria = limpiar(v_categoria)
        if not categoria:
            raise ValueError(f"'{hoja.title}' fila {fila}: categoría vacía (sector={sector_actual!r})")

        campo = f"pt_{sigla_actual}_{slug(categoria)}"
        contexto = f"'{hoja.title}' fila {fila} ({sector_actual} / {categoria})"

        planta.setdefault(sector_actual, {})
        registrar_en_mapa(planta[sector_actual], campo, categoria, vistos, contexto)
        fila += 1

    fila_fin_detectada = fila - 1
    if fila_fin_detectada != fila_fin_esperada:
        raise ValueError(
            f"'{hoja.title}': la tabla que empieza en {fila_ini} termina en la fila "
            f"{fila_fin_detectada} (total en {fila}), no en la fila {fila_fin_esperada} "
            "que se había verificado a mano. Revisa el instrumento antes de continuar."
        )
    return planta


# Un único diccionario de nombres vistos para las tres tablas: aunque los
# prefijos ('at_mc_', 'at_nat_', 'pt_') ya impiden choques entre bloques,
# comprobarlo de verdad -en vez de confiar en que los prefijos nunca
# cambien- es lo que exige el test que cuenta 113 campos únicos.
vistos = {}

hoja_atractivos = wb["ICT(Rt-At)"]
manifestaciones = leer_atractivos(
    hoja_atractivos, col_cat=2, col_tipo=3, col_subtipo=4,
    fila_ini=7, fila_fin_esperada=28, bloque_sigla="mc", vistos=vistos,
)
naturales = leer_atractivos(
    hoja_atractivos, col_cat=9, col_tipo=10, col_subtipo=11,
    fila_ini=7, fila_fin_esperada=61, bloque_sigla="nat", vistos=vistos,
)
atractivos = {**manifestaciones, **naturales}

hoja_planta = wb["ICT(Pt)"]
planta = leer_planta(
    hoja_planta, col_sector=10, col_categoria=11,
    fila_ini=8, fila_fin_esperada=43, vistos=vistos,
)


# manifestaciones/naturales son categoría -> tipo -> {campo: etiqueta}
# (dos niveles antes del mapa hoja); planta es sector -> {campo: etiqueta}
# (uno solo). Se cuenta cada forma explícitamente en vez de con una función
# genérica que adivine la profundidad del árbol.
n_mc = sum(len(mapa) for tipos in manifestaciones.values() for mapa in tipos.values())
n_nat = sum(len(mapa) for tipos in naturales.values() for mapa in tipos.values())
n_pt = sum(len(mapa) for mapa in planta.values())
print(f"// Manifestaciones culturales: {n_mc} subtipos")
print(f"// Atractivos naturales: {n_nat} subtipos")
print(f"// Atractivos (total): {n_mc + n_nat} subtipos")
print(f"// Planta turística: {n_pt} subcategorías en {len(planta)} sectores")
print(f"// Total general: {n_mc + n_nat + n_pt} campos, todos de {LIMITE_IDENTIFICADOR_POSTGRES} caracteres o menos")


def php_str(s):
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def php_mapa(mapa, indent):
    pad = " " * indent
    lineas = []
    for campo, etiqueta in mapa.items():
        lineas.append(f"{pad}{php_str(campo)} => {php_str(etiqueta)},")
    return lineas


lineas = [
    "<?php",
    "",
    "namespace App\\Matrices;",
    "",
    "/**",
    " * Índice de Concentración Turística.",
    " *",
    " * GENERADO por database/matrices/generar_concentracion.py",
    " * desde Documentación/Índice de Concentración Turística.xlsx.",
    " * No editar a mano: vuelve a ejecutar el generador.",
    " *",
    " * Instrumento de Calle Lituma y Chaca Espinoza, sobre Illingworth 2011. A",
    " * diferencia del resto de matrices del sistema, esta no puntúa criterios en",
    " * una escala: cuenta establecimientos y subtipos de atractivo por categoría.",
    " * No tiene escala acotada ni umbral: son conteos, no valoraciones.",
    " */",
    "final class Concentracion",
    "{",
    "    /**",
    "     * Atractivos turísticos: dos tablas paralelas del instrumento,",
    "     * manifestaciones culturales y atractivos naturales. El bloque de",
    "     * primer nivel es la categoría -la que separa las dos tablas y sus dos",
    "     * porcentajes-; dentro, agrupado por tipo (Arquitectura, Montaña,",
    "     * Folklore...), campo => etiqueta con el subtipo del instrumento.",
    "     *",
    "     * Algunos campos abrevian una palabra larga y recurrente del",
    "     * instrumento ('realizac' por 'realizaciones', 'aconteci' por",
    "     * 'acontecimientos'...) para no pasar el límite de 63 caracteres de un",
    "     * identificador de PostgreSQL -ver ABREVIATURAS en el generador-. Las",
    "     * etiquetas siempre llevan la palabra completa: la abreviatura es del",
    "     * nombre de columna, no del texto que ve el evaluador.",
    "     *",
    "     * @var array<string, array<string, array<string, string>>>",
    "     */",
    "    public const ATRACTIVOS = [",
]
for categoria, tipos in atractivos.items():
    lineas.append(f"        {php_str(categoria)} => [")
    for tipo, mapa in tipos.items():
        lineas.append(f"            {php_str(tipo)} => [")
        lineas += php_mapa(mapa, 16)
        lineas.append("            ],")
    lineas.append("        ],")
lineas += [
    "    ];",
    "",
    "    /**",
    "     * Planta turística: diez sectores del instrumento, cada uno con su",
    "     * sigla entre paréntesis en el nombre -es la misma que usan sus filas",
    "     * de subtotal en el instrumento y la que prefija sus campos-. Dentro de",
    "     * cada sector, campo => etiqueta con la subcategoría de establecimiento.",
    "     *",
    "     * @var array<string, array<string, string>>",
    "     */",
    "    public const PLANTA = [",
]
for sector, mapa in planta.items():
    lineas.append(f"        {php_str(sector)} => [")
    lineas += php_mapa(mapa, 12)
    lineas.append("        ],")
lineas += [
    "    ];",
    "",
    f"    /** @return array<int, string> los {n_mc + n_nat + n_pt} nombres de campo, en el orden del instrumento */",
    "    public static function campos(): array",
    "    {",
    "        $campos = [];",
    "        foreach (self::ATRACTIVOS as $tipos) {",
    "            foreach ($tipos as $mapa) {",
    "                $campos = array_merge($campos, array_keys($mapa));",
    "            }",
    "        }",
    "        foreach (self::PLANTA as $mapa) {",
    "            $campos = array_merge($campos, array_keys($mapa));",
    "        }",
    "",
    "        return $campos;",
    "    }",
    "}",
]

DESTINO.parent.mkdir(parents=True, exist_ok=True)
# newline="\n": sin esto, en Windows write_text traduce \n a \r\n y el
# fichero generado queda con finales de línea distintos a los del resto del
# repo (.gitattributes fija `eol=lf` para todo *.php).
DESTINO.write_text("\n".join(lineas) + "\n", encoding="utf-8", newline="\n")
print(f"escrito: {DESTINO}")
