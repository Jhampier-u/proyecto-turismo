"""Genera la definición PHP de los criterios de la Matriz de Valoración Territorial
directamente desde el instrumento original, para evitar errores de transcripción.

Uso, desde la raíz del proyecto:  python database/matrices/generar_valoracion_territorial.py
"""
import re
import openpyxl
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[2]
ORIGEN = RAIZ / "Documentación" / "Matriz de Valoración Territorial.xlsx"
DESTINO = RAIZ / "app" / "Matrices" / "ValoracionTerritorial.php"

# sigla en el instrumento -> nombre de columna en la base de datos
CAMPOS = {
    'EE': 'ct_energia_electrica',       'AP': 'ct_agua_potable',
    'SC': 'ct_comunicacion',            'RB': 'ct_recoleccion_basura',
    'PS': 'ct_problemas_sociales',      'SS': 'ct_salud',
    'SG': 'ct_seguridad',               'CR': 'ct_conservacion_recursos',
    'AE': 'ct_actividad_economica',     'OS': 'ct_organizacion_social',
    'DC': 'ct_elementos_culturales',    'DN': 'ct_espacios_naturales',
    'V':  'uc_vialidad',                'IC': 'uc_infraestructura_conectividad',
    'FC': 'uc_frecuencia_conectividad', 'DT': 'uc_distancia_atractivo',
    'DS': 'uc_distancia_sitio_visita',  'DD': 'uc_distancia_destino',
    'DM': 'uc_distancia_mercado_emisor','CO': 'uc_conglomeracion_oferta',
    'S':  'uc_senalizacion',
}

wb = openpyxl.load_workbook(ORIGEN, data_only=True)


def limpiar(t):
    return " ".join(str(t).split()) if t else ""


# Umbral de coincidencia usado por `mismo_criterio` (ver ese docstring para el
# porqué del valor). Con las 21 filas del instrumento, el par correcto peor
# alineado es CR y da 0.875 -no por diferencia terminológica, sino por una
# errata del instrumento ("reursos" en vez de "recursos", columna A fila 12
# de la hoja 'Valoración CT')-, y el peor cruce entre criterios distintos
# (SC/SS, que comparten el prefijo "Disponibilidad de Servicios de") da 0.75:
# 0.8 queda a medio camino, con margen a los dos lados.
UMBRAL_COINCIDENCIA = 0.8


def _palabras_clave(texto):
    """Conjunto de palabras (minúsculas, sin puntuación) de un nombre de
    criterio. Solo se usa para comparar hojas entre sí, nunca se vuelca al PHP."""
    texto = texto.lower()
    texto = re.sub(r"[^\wáéíóúñ]+", " ", texto)
    return {p for p in texto.split() if p}


def mismo_criterio(nombre_ref, nombre_val):
    """¿Describen `nombre_ref` (hoja de descripciones) y `nombre_val` (hoja de
    pesos, con la sigla al final entre paréntesis) el mismo criterio?

    No se compara por igualdad exacta porque el texto no es idéntico entre
    hojas: por ejemplo, para AP la hoja de pesos dice "...Agua Potable,
    Alcantarillado y Tratamiento de Aguas" y la de descripciones dice
    "...Agua Potable y Alcantarillado". Una igualdad estricta rompería el
    generador con el instrumento actual, que está bien alineado.

    Para eso se calcula el coeficiente de solapamiento de palabras:
    |palabras en común| / tamaño del conjunto más chico. Al dividir por el
    conjunto más chico, el lado con texto más largo no penaliza mientras sus
    palabras de más sean añadidos y no reemplazos: en AP, las 7 palabras de
    la hoja de descripciones están todas contenidas en las 9 de la hoja de
    pesos ("tratamiento" y "aguas" son las únicas que sobran), así que ese
    par da 1.0 pese a la diferencia de longitud entre hojas.

    El par correcto peor alineado del instrumento es CR, con 0.875 (7 de 8
    palabras coinciden). No es una discrepancia terminológica entre hojas:
    es una errata de tipeo en el instrumento. La hoja 'Valoración CT',
    columna A fila 12, dice "reursos" en vez de "recursos", así que esa
    palabra no coincide con la de la hoja de descripciones y queda fuera de
    la intersección.

    Se usa el coeficiente de solapamiento de palabras en vez de comparar las
    cadenas completas (con `difflib.SequenceMatcher.ratio`, por ejemplo)
    porque varios nombres comparten un prefijo largo tipo "Disponibilidad de
    Servicios de ..." y ese prefijo por sí solo ya empuja la similitud de
    cadena completa por encima del caso CR, borrando el margen entre pares
    correctos e incorrectos. Comparando conjuntos de palabras, ese prefijo
    compartido pesa lo mismo entre un par correcto y uno cruzado, y lo que
    decide es si la palabra distintiva (Comunicación, Salud, Alcantarillado...)
    también aparece en ambos lados.
    """
    sin_sigla = re.sub(r"\s*\([A-ZÁÉÍÓÚÑ]+\)\s*$", "", nombre_val).strip()
    palabras_ref, palabras_val = _palabras_clave(nombre_ref), _palabras_clave(sin_sigla)
    if not palabras_ref or not palabras_val:
        return False
    coincidencia = len(palabras_ref & palabras_val) / min(len(palabras_ref), len(palabras_val))
    return coincidencia >= UMBRAL_COINCIDENCIA


def leer(hoja_ref, fila_ini, fila_fin, hoja_val, val_ini, val_fin):
    """Empareja la hoja de descripciones con la de pesos. Ambas deben listar
    los criterios en el mismo orden: se comprueba que tengan el mismo número
    de filas y, fila a fila, que describan el mismo criterio (`mismo_criterio`)
    para detectar un desalineamiento silencioso si alguien reordena una hoja
    sin reordenar la otra."""
    ref, val = wb[hoja_ref], wb[hoja_val]
    filas_ref = list(range(fila_ini, fila_fin + 1))
    filas_val = list(range(val_ini, val_fin + 1))
    if len(filas_ref) != len(filas_val):
        raise ValueError(
            f"'{hoja_ref}' ({len(filas_ref)} filas) y '{hoja_val}' ({len(filas_val)} filas) "
            "no tienen el mismo número de criterios"
        )

    salida = []
    for fr, fv in zip(filas_ref, filas_val):
        nombre_ref = limpiar(ref.cell(fr, 1).value)     # columna A
        nombre_val = limpiar(val.cell(fv, 1).value)      # columna A
        if not mismo_criterio(nombre_ref, nombre_val):
            raise ValueError(
                f"desalineamiento entre '{hoja_ref}' fila {fr} y '{hoja_val}' fila {fv}: "
                f"'{nombre_ref}' no coincide con '{nombre_val}'"
            )
        sigla = limpiar(val.cell(fv, 6).value)          # columna F
        salida.append({
            'sigla': sigla,
            'campo': CAMPOS[sigla],
            'peso': val.cell(fv, 2).value,              # columna B
            'nombre': nombre_ref,
            'desc': [limpiar(ref.cell(fr, c).value) for c in (2, 3, 4)],  # B, C, D
        })
    return salida


ct = leer('Contenido Territorial', 6, 17, 'Valoración CT', 5, 16)
uc = leer('Ubicación y Conectividad', 6, 14, 'Valoración UC', 5, 13)

for nombre, grupo in (('CT', ct), ('UC', uc)):
    total = round(sum(c['peso'] for c in grupo), 10)
    if total != 1.0:
        raise ValueError(f"{nombre}: los pesos suman {total}, no 1.0")
    print(f"// {nombre}: {len(grupo)} criterios, pesos suman {total}")


def php_str(s):
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


lineas = ["<?php", "", "namespace App\\Matrices;", "",
          "/**", " * Criterios de la Matriz de Valoración Territorial.",
          " *", " * GENERADO por database/matrices/generar_valoracion_territorial.py",
          " * desde Documentación/Matriz de Valoración Territorial.xlsx.",
          " * No editar a mano: vuelve a ejecutar el generador.",
          " *", " * Instrumento de Calle Lituma y Chaca Espinoza. Los pesos de cada",
          " * dimensión suman 1 y la escala es 0-2, así que cada total va de 0 a 2.",
          " */", "class ValoracionTerritorial", "{",
          "    public const ESCALA_MIN = 0;",
          "    public const ESCALA_MAX = 2;", "",
          "    /** Umbral que separa los cuadrantes en ambos ejes. */",
          "    public const UMBRAL = 1.0;", ""]

for clave, grupo, titulo, fuente in (
    ('CT', ct, 'Contenido Territorial',
     'PDOT, PDT, fuentes secundarias y fuentes oficiales públicas. Para elementos '
     'culturales y espacios naturales: visitas in situ y documentos públicos.'),
    ('UC', uc, 'Ubicación y Conectividad',
     'PDOT, fuentes de información primaria y secundaria, visitas in situ y '
     'documentos oficiales.'),
):
    lineas.append(f"    /** {titulo}. Fuente sugerida: {fuente} */")
    lineas.append(f"    public const {clave} = [")
    for c in grupo:
        lineas.append(f"        {php_str(c['campo'])} => [")
        lineas.append(f"            'sigla'  => {php_str(c['sigla'])},")
        lineas.append(f"            'peso'   => {c['peso']},")
        lineas.append(f"            'nombre' => {php_str(c['nombre'])},")
        lineas.append("            'niveles' => [")
        for i, d in enumerate(c['desc']):
            lineas.append(f"                {i} => {php_str(d)},")
        lineas.append("            ],")
        lineas.append("        ],")
    lineas.append("    ];")
    lineas.append("")

lineas += ["    /** @return array<string, array> todos los criterios, CT seguidos de UC */",
           "    public static function todos(): array",
           "    {",
           "        return self::CT + self::UC;",
           "    }",
           "}"]

DESTINO.parent.mkdir(parents=True, exist_ok=True)
DESTINO.write_text("\n".join(lineas) + "\n", encoding="utf-8")
print(f"escrito: {DESTINO}")
