"""Genera la definición PHP de la Matriz de Análisis y Valoración del Paisaje
directamente desde el instrumento original, para evitar errores de transcripción.

Las etiquetas cualitativas de cada criterio no están en columnas limpias: viven
dentro de las fórmulas de la columna D, con la forma

    =IF(C11="Positivo","5",IF(C11="Neutro","3",IF(C11="Negativo","0")))

así que se extraen de ahí. Son 34 criterios con 3 etiquetas cada uno; copiarlas
a mano es justo donde se cuela el error que ningún test detecta.

Uso, desde la raíz del proyecto:  python database/matrices/generar_paisaje.py
"""
import re
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[2]
ORIGEN = RAIZ / "Documentación" / "Matriz de Análisis y Valoración del Paisaje.xlsx"
DESTINO = RAIZ / "app" / "Matrices" / "Paisaje.php"

# fila del instrumento -> nombre de columna en la base de datos.
# Se declaran explícitos en vez de derivarlos del nombre: un slug automático
# daría nombres frágiles que cambiarían si alguien corrige una tilde.
CAMPOS = {
    # Ep — Evolución del Paisaje
    11: 'ep_cambios_tiempo',        12: 'ep_imagen_pasada',
    13: 'ep_imagen_actual',         14: 'ep_imagen_futura',
    15: 'ep_conservacion',          16: 'ep_sostenibilidad',
    # Pn — Recursos Paisajísticos de Interés Natural
    18: 'pn_areas_verdes',          19: 'pn_areas_protegidas',
    20: 'pn_geomorfologia',         21: 'pn_topografia',
    22: 'pn_hidrografia',           23: 'pn_geologia',
    24: 'pn_flora',                 25: 'pn_fauna',
    26: 'pn_cobertura_suelo',
    # Pc — Recursos Paisajísticos de Interés Cultural
    28: 'pc_arquitectonico_moderno', 29: 'pc_vernacular',
    30: 'pc_edificado_historico',    31: 'pc_religioso',
    32: 'pc_vivo',
    # Iv — Recursos Paisajísticos de Interés Visual
    34: 'iv_referentes_naturales',  35: 'iv_referentes_antropicos',
    36: 'iv_puntos_observacion',    37: 'iv_singularidad',
    38: 'iv_representatividad',     39: 'iv_visibilidad',
    40: 'iv_calidad_escenica',
    # Cp — Conflictos Paisajísticos
    42: 'cp_degradacion_natural',   43: 'cp_degradacion_antropica',
    44: 'cp_dinamicas_territoriales', 45: 'cp_fragmentacion',
    46: 'cp_infraestructura',       47: 'cp_actividades_diversas',
    48: 'cp_conurbaciones',
}

# sigla -> (nombre de la categoría, primera fila, última fila, fila del peso)
CATEGORIAS = [
    ('ep', 'Evolución del Paisaje',                        11, 16, 34),
    ('pn', 'Recursos Paisajísticos de Interés Natural',    18, 26, 35),
    ('pc', 'Recursos Paisajísticos de Interés Cultural',   28, 32, 36),
    ('iv', 'Recursos Paisajísticos de Interés Visual',     34, 40, 37),
    ('cp', 'Conflictos Paisajísticos',                     42, 48, 38),
]

# Escenarios del resultado final: filas 34-38, columnas U (nombre) y W (texto).
FILAS_ESCENARIO = range(34, 39)

wb_val = openpyxl.load_workbook(ORIGEN, data_only=True)
wb_frm = openpyxl.load_workbook(ORIGEN, data_only=False)
hv, hf = wb_val['PRT'], wb_frm['PRT']


def limpiar(t):
    return " ".join(str(t).split()) if t is not None else ""


def niveles_de(fila):
    """Extrae etiqueta -> valor de la fórmula IF anidada de la columna D."""
    formula = hf.cell(fila, 4).value or ""
    pares = re.findall(r'"([^"]+)"\s*,\s*"(\d+)"', formula)

    if len(pares) != 3:
        raise ValueError(
            f"fila {fila}: se esperaban 3 niveles en la fórmula, se hallaron {len(pares)}: {formula!r}"
        )

    return {int(valor): etiqueta for etiqueta, valor in pares}


def php(s):
    return "'" + str(s).replace("\\", "\\\\").replace("'", "\\'") + "'"


categorias = []

for sigla, nombre, fila_ini, fila_fin, fila_peso in CATEGORIAS:
    peso = hv.cell(fila_peso, 17).value  # columna Q

    criterios = []
    for fila in range(fila_ini, fila_fin + 1):
        if fila not in CAMPOS:
            raise ValueError(f"fila {fila} de {sigla} sin nombre de campo asignado")

        criterios.append({
            'campo': CAMPOS[fila],
            'nombre': limpiar(hv.cell(fila, 2).value),   # columna B
            'niveles': niveles_de(fila),
        })

    categorias.append({
        'clave': sigla, 'nombre': nombre, 'peso': peso, 'criterios': criterios,
    })

total_pesos = round(sum(c['peso'] for c in categorias), 10)
if total_pesos != 1.0:
    raise ValueError(f"los pesos de las categorías suman {total_pesos}, no 1.0")

total_criterios = sum(len(c['criterios']) for c in categorias)
if total_criterios != len(CAMPOS):
    raise ValueError(f"se recorrieron {total_criterios} criterios y hay {len(CAMPOS)} campos declarados")

escenarios = []
for fila in FILAS_ESCENARIO:
    escenarios.append({
        'rango': limpiar(hv.cell(fila, 20).value),   # columna T
        'nombre': limpiar(hv.cell(fila, 21).value),  # columna U
        'texto': limpiar(hv.cell(fila, 23).value),   # columna W
    })

for e in escenarios:
    if not (e['nombre'] and e['texto']):
        raise ValueError(f"escenario incompleto: {e}")

for c in categorias:
    print(f"// {c['clave'].upper()}: {len(c['criterios'])} criterios, peso {c['peso']}")
print(f"// total: {total_criterios} criterios, pesos suman {total_pesos}")
print(f"// escenarios: {len(escenarios)}")

L = ["<?php", "", "namespace App\\Matrices;", "",
     "/**", " * Criterios de la Matriz de Análisis y Valoración del Paisaje.",
     " *", " * GENERADO por database/matrices/generar_paisaje.py desde",
     " * Documentación/Matriz de Análisis y Valoración del Paisaje.xlsx.",
     " * No editar a mano: vuelve a ejecutar el generador.",
     " *", " * Instrumento de Calle Lituma y Chaca Espinoza, sobre Muñoz 2012 y",
     " * Bertinni 2019. Cada criterio se califica 0, 3 o 5 con etiquetas propias;",
     " * cada categoría promedia los suyos y el promedio se pondera. Los pesos",
     " * suman 1, así que el resultado final va de 0 a 5.", " */",
     "class Paisaje", "{",
     "    /** Valores admitidos por el instrumento. No es un rango continuo. */",
     "    public const VALORES = [0, 3, 5];", "",
     "    public const MAXIMO = 5;", ""]

L.append("    /** @var array<string, array> categoría => nombre, peso y criterios */")
L.append("    public const CATEGORIAS = [")
for c in categorias:
    L.append(f"        {php(c['clave'])} => [")
    L.append(f"            'nombre' => {php(c['nombre'])},")
    L.append(f"            'peso'   => {c['peso']},")
    L.append("            'criterios' => [")
    for cr in c['criterios']:
        L.append(f"                {php(cr['campo'])} => [")
        L.append(f"                    'nombre'  => {php(cr['nombre'])},")
        L.append("                    'niveles' => [")
        for valor in sorted(cr['niveles']):
            L.append(f"                        {valor} => {php(cr['niveles'][valor])},")
        L.append("                    ],")
        L.append("                ],")
    L.append("            ],")
    L.append("        ],")
L.append("    ];")
L.append("")

L.append("    /** Escenarios del resultado final, de mayor a menor. */")
L.append("    public const ESCENARIOS = [")
for e in escenarios:
    L.append("        [")
    L.append(f"            'rango'  => {php(e['rango'])},")
    L.append(f"            'nombre' => {php(e['nombre'])},")
    L.append(f"            'texto'  => {php(e['texto'])},")
    L.append("        ],")
L.append("    ];")
L.append("")

L += ["    /** @return array<string, array> los 34 criterios, aplanados */",
      "    public static function todos(): array",
      "    {",
      "        $criterios = [];",
      "",
      "        foreach (self::CATEGORIAS as $categoria) {",
      "            $criterios += $categoria['criterios'];",
      "        }",
      "",
      "        return $criterios;",
      "    }",
      "}"]

DESTINO.parent.mkdir(parents=True, exist_ok=True)
DESTINO.write_text("\n".join(L) + "\n", encoding="utf-8")
print(f"escrito: {DESTINO}")
